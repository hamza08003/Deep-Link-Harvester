import os
import sys
import json
import asyncio
from typing import Optional, Dict, Any
from collections import deque, defaultdict
from datetime import datetime, timedelta
from urllib.parse import urlparse, urljoin, parse_qs, urlencode

import tldextract
import pandas as pd
from rich import box
from rich.table import Table
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


from crawl4ai import (
    AsyncWebCrawler, BrowserConfig, CrawlerRunConfig, RateLimiter,
    MemoryAdaptiveDispatcher, CrawlerMonitor, DisplayMode, CacheMode
)
from crawl4ai.models import CrawlStats, CrawlStatus



# Media Extensions to Skip if they are found in the URL path
MEDIA_EXTENSIONS_TO_SKIP = [
    ".jpg", ".jpeg", ".png", ".gif", ".svg", ".mp4", ".mov", ".webm",
    ".ogv", ".flv", ".mfv", ".m4v", ".mpg", ".mpeg", ".mkv", ".wmv",
    ".wav", ".mp3", ".aac", ".flac", ".ogg", ".weba", ".opus", ".ico",
    ".bmp", ".tiff", ".tif", ".webp", ".mpd", ".pdf", ".psd", ".raw",
    ".cr2", ".nef", ".orf", ".sr2", ".arw"
]

# Keywords to skip if found in the URL path
PROHIBITED_PATH_KEYWORDS = {
    "image", "img", "icon", "video", "audio", "file", 
    "login", "log-in", "register", "session", "account", "auth",
    "logout", "log-out", "signin", "sign-in", "preferences", "settings",
    "profile", "id", "token", "form", "submit", "user",
    "password", "security", "secure", "payment", "checkout",
}


def get_main_domain(domain_url: str) -> str:
    """
    This function parses the provided URL to extract its domain and public suffix.

    For example:
    - https://mail.google.com/mail/u/0/'       -> 'google.com'
    - https://www.speedtest.net/apps/windows   -> 'speedtest.net'
    - http://forums.bbc.co.uk/                 -> 'bbc.co.uk'
    - ftp://ftp.debian.org/debian/             -> 'debian.org'

    Args:
        domain_url (str): The URL from which to extract the domain and suffix.

    Returns:
        str: A string in the format 'domain.suffix' if a valid suffix is found;
             otherwise, the original netloc of the URL.
    """
    try:
        # break down the given url into its components: scheme, netloc, path, params, query, and fragment.
        parsed = urlparse(domain_url)
        # extract the subdomain, domain, and suffix from the netloc
        ext = tldextract.extract(parsed.netloc)
        # return the domain and suffix if a valid suffix is found, otherwise return the original netloc
        return f"{ext.domain}.{ext.suffix}" if ext.suffix else parsed.netloc
    except:
        return parsed.netloc


def normalize_url(url: str, base_url: str) -> Optional[str]:
    """
    Normalize a URL to a standard format, handling relative URLs, skipping media URLs and the URLs containing prohibited keywords.

    Args:
        url (str): The URL to normalize.
        base_url (str): The base URL to use for relative URLs.

    Returns:
        `str`: The normalized URL.
    """

    try:
        # break down the given url into its components: scheme, netloc, path, params, query, and fragment.
        # for example, if the URL is https://www.example.com/page?name=John&age=30&page=1#section-1
        # the parsed URL will be Object of type `ParseResult` and will have the following attributes:
        # scheme='https', netloc='www.example.com', path='/page', params='', query='name=John&age=30&page=1', fragment='section-1'
        parsed = urlparse(url)

        # check if the parsed URL does NOT have a network location (i.e., it's a relative URL)
        if not parsed.netloc:
            # combine the base_url with the relative path to form a complete URL.
            full_url = urljoin(base_url, parsed.path)
            # re-parse the full URL to get the components
            parsed = urlparse(full_url)
        
        # get the path of the URL and convert it to lowercase
        path = parsed.path.lower()

        # check if the URLs ends with any of the media extensions to skip or contains any of the prohibited keywords
        if any(path.endswith(ext) for ext in MEDIA_EXTENSIONS_TO_SKIP) or \
           any(keyword in path.split('/') for keyword in PROHIBITED_PATH_KEYWORDS):
            return None  # skip media URLs and URLs containing prohibited keywords
        
        # parse the query string (everything after ?) into a dictionary,
        # where keys are query parameter names and values are lists of values.
        # for example, if the URL is https://www.example.com/page?name=John&age=30&page=1,
        # the query will be {'name': ['John'], 'age': ['30'], 'page': ['1']}
        query = parse_qs(parsed.query)

        # sort the query parameters by their names and encode them into a URL-encoded string,
        # except for the 'page' parameter, taking the above example,
        # keys are converted to lowercase and sorted, so the dictionary will be {'age': ['30'], 'name': ['John']}
        # and then finally the sorted query will be 'age=30&name=John'
        sorted_query = urlencode({k.lower(): sorted(v) for k, v in query.items() if k.lower() not in ['page']}, doseq=True)
        
        # replace the parsed URL components with the normalized ones
        normalized = parsed._replace(
            scheme=parsed.scheme.lower(),   # convert scheme (http/https) to lowercase
            netloc=parsed.netloc.lower(),   # convert netloc (domain name) to lowercase
            path=parsed.path.lower(),       # convert path to lowercase
            query=sorted_query,             # use the cleaned query string
            fragment=""                     # clear the fragment part
        ).geturl().rstrip('/')              # convert the parsed URL back to a string and remove any trailing slashes
        return normalized
    
    except Exception as e:
        print(f"Error normalizing URL: {e}")
        return None
    

class CustomCrawlerConfig(CrawlerRunConfig):
    """
    Custom Crawler Config Override from Crawl4AI
    """
    def __init__(
        self,
        max_depth: int = 5,
        urls_per_depth: list[str | int] = None,
        **kwargs
    ):
        super().__init__(**kwargs)
        self.max_depth = max_depth
        self.urls_per_depth = urls_per_depth or ['all'] + ['all'] * max_depth  # default to 'all' for each level


class CustomCrawlerMonitor(CrawlerMonitor):
    """
    Custom monitor that extends CrawlerMonitor with enhanced status messages
    """
    
    def _create_detailed_table(self) -> Table:
        table = Table(
            box=box.ROUNDED,
            title="Crawler Performance Monitor",
            title_style="bold magenta",
            header_style="bold blue",
        )

        # add columns
        table.add_column("Task ID", style="cyan", no_wrap=True)
        table.add_column("URL", style="cyan", no_wrap=True)
        table.add_column("Status", style="bold")
        table.add_column("Memory (MB)", justify="right")
        table.add_column("Peak (MB)", justify="right")
        table.add_column("Duration", justify="right")
        table.add_column("Info", style="italic")

        # add summary row with enhanced formatting
        total_memory = sum(stat.memory_usage for stat in self.stats.values())
        active_count = sum(1 for stat in self.stats.values() if stat.status == CrawlStatus.IN_PROGRESS)
        completed_count = sum(1 for stat in self.stats.values() if stat.status == CrawlStatus.COMPLETED)
        failed_count = sum(1 for stat in self.stats.values() if stat.status == CrawlStatus.FAILED)

        table.add_row(
            "[bold yellow]SUMMARY",
            f"Total: {len(self.stats)}",
            f"Active: {active_count}",
            f"{total_memory:.1f}",
            f"{self.process.memory_info().rss / (1024 * 1024):.1f}",
            str(timedelta(seconds=int((datetime.now() - self.start_time).total_seconds()))),
            f"[green]✓ {completed_count}[/green] [red]✗  {failed_count}[/red]",
            style="bold",
        )

        table.add_section()

        # add rows for each task with enhanced status messages
        visible_stats = sorted(
            self.stats.values(),
            key=lambda x: (
                x.status != CrawlStatus.IN_PROGRESS,
                x.status != CrawlStatus.QUEUED,
                x.end_time or datetime.max,
            ),
        )[: self.max_visible_rows]

        for stat in visible_stats:
            status_style = {
                CrawlStatus.QUEUED: "white",
                CrawlStatus.IN_PROGRESS: "yellow",
                CrawlStatus.COMPLETED: "green",
                CrawlStatus.FAILED: "red",
            }[stat.status]

            # enhanced status messages with icons and details
            status_message = self._get_enhanced_status_message(stat)

            table.add_row(
                stat.task_id[:8],
                stat.url[:40] + "..." if len(stat.url) > 40 else stat.url,
                f"[{status_style}]{stat.status.value}[/{status_style}]",
                f"{stat.memory_usage:.1f}",
                f"{stat.peak_memory:.1f}",
                stat.duration,
                status_message,
            )

        return table

    def _get_enhanced_status_message(self, stat: CrawlStats) -> str:
        """
        Generate enhanced status messages with icons and details
        """
        if stat.status == CrawlStatus.COMPLETED:
            return "[green]✓ Successfully Scraped[/green]"
        elif stat.status == CrawlStatus.FAILED:
            error_msg = stat.error_message[:40] if stat.error_message else "Unknown error"
            return f"[red]✗ Failed: {error_msg}[/red]"
        elif stat.status == CrawlStatus.IN_PROGRESS:
            return "[yellow]⟳ Processing URL...[/yellow]"
        elif stat.status == CrawlStatus.QUEUED:
            return "[white]⋯ In Queue[/white]"
        return ""
    

def load_config(config_file: str) -> Dict[str, Any]:
    """
    Load and validate configuration from JSON file

    Args:
        config_file (str): The path to the JSON file containing the configuration

    Returns:
        `dict`: A dictionary containing the configuration
    """
    try:
        with open(config_file, 'r') as f:
            config = json.load(f)
        
        # validate basic structure
        if not isinstance(config, dict):
            raise ValueError("Config must be a JSON object")
        
        if ('browser_config' not in config or 
            'crawler_config' not in config or 
            'other_config' not in config):
            raise ValueError("Config must contain 'browser_config', 'crawler_config', and 'other_config' sections")
            
        return config
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON format: {str(e)}")
    except Exception as e:
        raise ValueError(f"Error loading config: {str(e)}")


def create_browser_config(config: Dict[str, Any]) -> BrowserConfig:
    """
    Create a `BrowserConfig` Object from given dictionary.

    Args:
        config (dict): A dictionary containing the browser configuration

    Returns:
        `BrowserConfig`: A BrowserConfig object
    """
    try:
        return BrowserConfig(**config)
    except Exception as e:
        raise ValueError(f"Invalid browser configuration: {str(e)}")


def create_crawler_config(config: Dict[str, Any], max_depth: int, urls_per_depth: list) -> CustomCrawlerConfig:
    """
    Create `CustomCrawlerConfig` from given dictionary.

    Args:
        config (dict): A dictionary containing the crawler configuration
        max_depth (int): The maximum depth level
        urls_per_depth (list): The number of URLs to crawl per depth level

    Returns:
        `CustomCrawlerConfig`: A CustomCrawlerConfig object
    """
    try:
        # handle special case for cache_mode if it's a string
        if 'cache_mode' in config and isinstance(config['cache_mode'], str):
            config['cache_mode'] = CacheMode[config['cache_mode']]
        
        # add runtime crawler params to the CustomCrawlerConfig via kwargs
        return CustomCrawlerConfig(
            max_depth=max_depth,
            urls_per_depth=urls_per_depth,
            **config
        )
        
    except Exception as e:
        raise ValueError(f"Invalid crawler configuration: {str(e)}")


async def load_domains(domains_file: str) -> list[str]:
    """
    Loads a list of domain URLs from an Excel file

    Args:
        domains_file (str): The path to the Excel file containing the domains to crawl

    Returns:
        `list[str]`: A list of domain URLs
    """
    df = pd.read_excel(domains_file)
    return [url.strip() for url in df['Domain'].tolist() if isinstance(url, str)]


async def save_results(results: dict, filename: str) -> None:
    """
    Saves the Crawling Results to an Excel file.

    Args:
        results (dict): A dictionary containing the crawling results with the following structure:
            {
                "domain": str,
                "total_links": int,
                "urls_by_depth": dict,
                "images": list
            }
        Here 'urls_by_depth' is a dictionary with depth levels as keys and lists of URLs as values.
        And 'images' is a list of tuples containing image address URLs, page source URLs, and timestamps.

        filename (str): The name of the output Excel file to save the results
    
    Returns:
        None
    """

    # Create a new Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sort depths to create sheets in order
    depths = sorted(results['urls_by_depth'].keys())
    
    # Create lookup dictionary for images by page URL
    images_by_page = defaultdict(list)
    if 'images' in results and results['images']:
        for image_url, page_url, timestamp in results['images']:
            images_by_page[page_url].append((image_url, timestamp))
    
    # Track image counts per depth level
    image_counts_by_depth = defaultdict(int)
    
    # Loop through each depth level to Create Seperate Sheets
    for depth in depths:
        sheet_name = f"Depth Level {depth}"
        ws = wb.create_sheet(sheet_name)
        
        # Header Row
        ws.append(["Page URL (Image Source)", "Image Address", "Timestamp"])
        
        # Style the header cells with light blue fill, font size 16, bold, and centered text
        ds_header_fill = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = Font(size=16, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = ds_header_fill
        
        # Process URLs for this depth   
        for url, page_timestamp in results['urls_by_depth'][depth]:
            page_images = images_by_page.get(url, [])
            
            if not page_images:
                # If page has no images, add just the page URL row
                ws.append([url, "", page_timestamp])
            else:
                # # for the first image, include the page URL
                # first_image = page_images[0]
                # ws.append([url, first_image[0], first_image[1]])
                # image_counts_by_depth[depth] += 1
                
                # # for additional images, leave page URL blank for clearer grouping
                # for image_url, image_timestamp in page_images[1:]:
                #     ws.append(["", image_url, image_timestamp])
                #     image_counts_by_depth[depth] += 1

                # Include page URL for EVERY image (not just the first one)
                for image_url, image_timestamp in page_images:
                    ws.append([url, image_url, image_timestamp])
                    image_counts_by_depth[depth] += 1
        
        # set remaining rows font size = 12, center align
        for row in ws.iter_rows(min_row=2):  # skip header
            for cell in row:
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # auto-adjust column widths based on content
        for idx, col in enumerate(ws.columns):
            if idx == 0:
                # # adjust width of first column (Page URL) based on longest URL
                # max_length = max(len(str(cell.value)) for cell in col)
                # ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5
                
                # custom width for Page URL column
                ws.column_dimensions[get_column_letter(col[0].column)].width = 100

            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            ws.column_dimensions[col_letter].width = max_length + 5

    # Create a Summary sheet
    summary = wb.create_sheet("Summary", 0)
    
    # --- Top row: Domain (A1) and domain value (B1) ---
    summ_header_fill = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")
    
    summary["A1"] = "Domain"
    summary["B1"] = results['domain']
    
    # style for A1
    summary["A1"].font = Font(color="FFB70016", size=16, bold=True)  # Red
    summary["A1"].fill = summ_header_fill
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # style for B1
    summary["B1"].font = Font(color="FF00008B", size=16, bold=True)  # Dark Blue
    summary["B1"].fill = summ_header_fill
    summary["B1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # --- Next rows: Total URLs, Total Images ---
    summary.append(["Total URLs", results['total_links']])
    
    # styling for Total URLs row
    summary["A2"].font = Font(color="3973AC", size=14, bold=True)  # for 'Total URLs' label
    summary["A2"].alignment = Alignment(horizontal="center", vertical="center")
    summary["B2"].font = Font(color="008000", size=14, bold=True)  # for URL count
    summary["B2"].alignment = Alignment(horizontal="center", vertical="center")

    if 'images' in results:
        summary.append(["Total Images", len(results['images'])])
        # styling for Total Images row
        summary["A3"].font = Font(color="3973AC", size=14, bold=True)  # for 'Total Images' label
        summary["A3"].alignment = Alignment(horizontal="center", vertical="center")
        summary["B3"].font = Font(color="A020F0", size=14, bold=True)  # for image count
        summary["B3"].alignment = Alignment(horizontal="center", vertical="center")

    summary.append([])  # blank row
    
    # --- Table header: Depth Level, Number of URLs, Number of Images ---
    summary.append(["Depth Level", "Number of URLs", "Number of Images"])
    header_row_idx = summary.max_row
    for col_idx, cell in enumerate(summary[header_row_idx], start=1):
        cell.font = Font(size=14, bold=True)
        cell.fill = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")  # light blue fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # --- Depth data rows ---
    for depth in depths:
        summary.append([
            f"Depth {depth}",
            len(results['urls_by_depth'][depth]),
            image_counts_by_depth[depth]
        ])
    
        # Apply styling to the counts in the current row
        row_idx = summary.max_row

        row_idx = summary.max_row
        summary.cell(row=row_idx, column=1).font = Font(color="800000", size=13, bold=True)  # Depth Level (maroon)
        summary.cell(row=row_idx, column=2).font = Font(color="008000", size=13, bold=True)  # URLs (green)
        summary.cell(row=row_idx, column=3).font = Font(color="FFA500", size=13, bold=True)  # Images (orange)

        for col_idx in range(1, 4):
            summary.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
    
    # auto-adjust column widths on Summary
    for col in summary.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        # add a little padding
        summary.column_dimensions[col_letter].width = max_length + 10
    
    # Save the workbook
    wb.save(filename)
    print(f"\nResults saved to {filename}")
    
    # Print image statistics
    total_images = sum(image_counts_by_depth.values())
    print(f"Total images found: {total_images}")
    for depth, count in image_counts_by_depth.items():
        print(f"Images at depth {depth}: {count}")


async def harvest_links(domain_url: str, browser_config: BrowserConfig, run_config: CustomCrawlerConfig, depth_wise_url_batch_size: int) -> None:
    """
        Crawls a domain recursively, up to the specified depth and URLs per depth level,
        processing URLs in batches to manage memory and collecting images

        Args:
            domain_url (str): The URL of the domain to crawl
            browser_config (BrowserConfig): The browser configuration object
            run_config (CustomCrawlerConfig): The crawler runtime configuration object
            depth_wise_url_batch_size (int): The number of URLs to process in each batch per depth

        Returns:
            None
    """
    
    main_domain = get_main_domain(domain_url)
    crawled_pages = 0
    max_depth = run_config.max_depth
    visited = set()
    to_visit = deque([(domain_url, 0)])
    urls_by_depth = defaultdict(list)  # store URLs grouped by depth
    all_images = []  # storage for images with their page URLs
    depth_batch_results = defaultdict(list)  # temporary storage for batch results
    
    async with AsyncWebCrawler(config=browser_config) as crawler:
        while to_visit:
            current_depth = to_visit[0][1] if to_visit else max_depth + 1
            if current_depth > max_depth:
                break
            
            print(f"\n{'='*100}")
            print(f"Processing Depth Level {current_depth}")
            print(f"{'='*100}\n")
            
            # Collect all URLs for current depth
            current_level_urls = []
            depth = current_depth
            while to_visit and to_visit[0][1] == depth:
                url, _ = to_visit.popleft()
                current_level_urls.append(url)
            
            total_at_depth = len(current_level_urls)
            
            # Apply URLs per depth limit
            urls_limit = run_config.urls_per_depth[depth]
            if urls_limit != 'all' and isinstance(urls_limit, int):
                current_level_urls = current_level_urls[:urls_limit]
            
            print(f"Found {total_at_depth} URLs at depth {depth} (limit: {urls_limit})")
            
            # Process URLs in batches
            for batch_start in range(0, len(current_level_urls), depth_wise_url_batch_size):
                batch_urls = current_level_urls[batch_start:batch_start + depth_wise_url_batch_size]
                batch_num = (batch_start // depth_wise_url_batch_size) + 1
                total_batches = (len(current_level_urls) + depth_wise_url_batch_size - 1) // depth_wise_url_batch_size
                
                print(f"\nProcessing batch {batch_num}/{total_batches} at depth {depth}")
                print(f"Batch size: {len(batch_urls)} URLs")
                
                # create fresh monitor for each batch
                monitor = CustomCrawlerMonitor(
                    display_mode=DisplayMode.DETAILED,
                    max_visible_rows=15
                )
                dispatcher = MemoryAdaptiveDispatcher(
                    memory_threshold_percent=75.0, # memory threshold percentage
                    max_session_permit=10, # maximum number of concurrent sessions
                    rate_limiter=RateLimiter(
                        base_delay=(1.5, 3.0), # base delay for rate limiting
                        max_delay=30.0, # maximum delay for rate limiting
                        max_retries=3, # maximum number of retries for rate limiting
                        rate_limit_codes=[400, 401, 403, 404, 408, 420, 421, 425, 429, 500, 502, 503, 504, 509] # status codes to trigger rate limiting
                    ),
                    monitor=monitor
                )
                
                # Crawl the current batch
                try:
                    results = await crawler.arun_many(
                        urls=batch_urls,
                        config=run_config,
                        dispatcher=dispatcher
                    )
                except Exception as e:
                    print(f"\nError crawling batch at depth {depth}: {str(e)}")
                    continue
                
                # Process batch results
                next_level_urls = []
                for url, result in zip(batch_urls, results):
                    if result.success:
                        crawled_pages += 1
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        visited.add((url, depth, timestamp))
                        depth_batch_results[depth].append((url, timestamp))
                        
                        # Extract and store image information
                        if hasattr(result, 'media') and 'images' in result.media:
                            for image in result.media['images']:
                                if 'src' in image and image['src']:
                                    # ensure image URL is absolute
                                    image_url = image['src']
                                    
                                    # convert relative URLs to absolute URLs
                                    if image_url.startswith('/'):
                                        # get base URL (scheme + domain)
                                        try:
                                            page_url_obj = urlparse(url)
                                            base_url = f"{page_url_obj.scheme}://{page_url_obj.netloc}"
                                            image_url = urljoin(base_url, image_url)
                                        except Exception:
                                            # keep original if parsing fails
                                            pass
                                    elif not image_url.startswith(('http://', 'https://')):
                                        # handle other relative formats
                                        image_url = urljoin(url, image_url)
                                    
                                    # store as tuple: (image_url, page_url, timestamp)
                                    all_images.append((image_url, url, timestamp))
                        
                        # Extract Internal Links (Page Source URLs) for next level
                        internal_links = result.links.get("internal", [])
                        for link in internal_links:
                            # get the href attribute from the link
                            href = link.get("href", "")
                            # normalize the URL, check if it's not already visited and add to next level URLs list
                            normalized = normalize_url(href, result.url)
                            if normalized and normalized not in visited:
                                visited.add(normalized)
                                # store as tuple: (normalized_url, depth)
                                next_level_urls.append((normalized, depth + 1)) 
                
                # Add next level URLs to the queue
                to_visit.extend(next_level_urls)
                
                # Clear some memory after each batch
                del results
                
            # After all batches for current depth are processed, merge the results and Clear the batch results
            urls_by_depth[depth].extend(depth_batch_results[depth])
            depth_batch_results[depth].clear()
            
        print(f"\nCrawl completed. Found {len(all_images)} images from {crawled_pages} pages.")
        
        # Save final results
        await save_results(
            {
                "domain": main_domain,
                "total_links": len(visited),
                "urls_by_depth": urls_by_depth,
                "images": all_images
            },
            f"{main_domain}_complete_data.xlsx"
        )
        
        print(f"\nCrawl completed for {domain_url}. Total crawled pages: {crawled_pages} | Total unique links: {len(visited)} | Total images: {len(all_images)}\n")


async def main(domains_file: str, browser_config: BrowserConfig, run_config: CustomCrawlerConfig, depth_wise_url_batch_size: int) -> None:
    """
    Main function to orchestrate crawling of multiple domains from an Excel file

    Args:
        domains_filepath (str): The path to the Excel file containing the domains to crawl
        browser_config (BrowserConfig): The browser configuration for the crawler
        run_config (CustomCrawlerConfig): The run configuration for the crawler

    Returns:
        None
    """
    domains = await load_domains(domains_file)
    total_domains = len(domains)
    print(f"\nTotal domains to process: {total_domains}\n{'-' * 100}")

    for index, domain in enumerate(domains, 1):
        print(f"\nProcessing domain {index}/{total_domains}: {domain}")
        try:
            await harvest_links(domain, browser_config, run_config, depth_wise_url_batch_size)
        except Exception as e:
            print(f"\nError processing domain {domain}: {str(e)}")
        print(f"\nDomain {index}/{total_domains} processed\n{'-' * 100}")


if __name__ == "__main__":
    try:
        print("\nImage URLs with Page Source URLs - Harvester v2.0")
        print("=" * 50)
        
        # Get Excel file path from user
        while True:
            domains_file = input("\nEnter the path to your domains Excel file: ").strip('"').strip("'")
            if os.path.exists(domains_file) and domains_file.lower().endswith('.xlsx'):
                break
            print("Error: Invalid file path or not an Excel file. Please try again.")

        # Get config file path from user
        while True:
            config_file = input("\nEnter the path to your config JSON file: ").strip('"').strip("'")
            if os.path.exists(config_file) and config_file.lower().endswith('.json'):
                try:
                    config = load_config(config_file)
                    browser_conf = create_browser_config(config['browser_config'])
                    break
                except ValueError as e:
                    print(f"Error in config file: {str(e)}")
                    continue
            print("Error: Invalid file path or not a JSON file. Please try again.")

        # Get max_depth from user
        while True:
            try:
                max_depth = int(input("\nEnter the maximum depth level: "))
                if max_depth > 0:
                    break
                print("Please enter a positive number")
            except ValueError:
                print("Please enter a valid number")

        # Get urls_per_depth from user
        urls_per_depth = ['all']  # depth 0 is always 'all'
        
        print("\n🔍 For each depth level (starting from depth 1; depth level 0 is always set to 'all'), specify the number of URLs to crawl:")
        print("✅ Type 'all' or press Enter to crawl all the URLs at that depth level.")
        print("🔢 Or enter a specific number to limit the URLs crawled at that depth.")
        
        for depth in range(1, max_depth + 1):
            while True:
                urls_input = input(f"Enter URLs limit for depth {depth} (default: all): ").lower()
                if not urls_input:  # if user just pressed Enter
                    urls_per_depth.append('all')
                    break
                if urls_input == 'all':
                    urls_per_depth.append('all')
                    break
                try:
                    urls_limit = int(urls_input)
                    if urls_limit > 0:
                        urls_per_depth.append(urls_limit)
                        break
                    print("Please enter a positive number, 'all', or press Enter")
                except ValueError:
                    print("Please enter a valid number, 'all', or press Enter")

        # create crawler config with runtime parameters
        run_conf = create_crawler_config(config['crawler_config'], max_depth, urls_per_depth)

        # get batch_size from the config, default to 500 if not present
        depth_wise_url_batch_size = config.get('other_config', {}).get('depth_wise_url_batch_size', 500)
        if not isinstance(depth_wise_url_batch_size, int) or depth_wise_url_batch_size <= 0:
            print(f"Warning: Invalid depth_wise_url_batch_size ({depth_wise_url_batch_size}), using default of 500")
            batch_size = 500

        # run the crawler
        asyncio.run(main(domains_file, browser_conf, run_conf, depth_wise_url_batch_size))
        
    except Exception as e:
        print(f"\nAn error occurred: {str(e)}")
    finally:
        input("\nPress Enter to exit...")
